import json
import logging
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import requests
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.repositories.fee_config_repository import fee_config_repo
from app.repositories.invoice_repository import invoice_repo
from app.repositories.port_log_repository import port_log_repo
from app.repositories.export_log_repository import export_log_repo
from app.services.storage.minio_service import get_object_bytes, parse_minio_uri, presign_get
from app.utils.fee_billing_unit import fee_billing_unit_label

EXPORTS_DIR = Path('exports')
_HEADER_FILL = PatternFill('solid', fgColor='5B9BD5')
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_ACCOUNTING_HEADER_FILL = PatternFill('solid', fgColor='FFF200')
_ACCOUNTING_HEADER_FONT = Font(bold=True, color='000000')
_THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)
_IMAGE_FETCH_TIMEOUT_SEC = 15
_IMAGE_THUMBNAIL_SIZE_PX = (160, 90)
_log = logging.getLogger('app.services.export')


class ExportService:
    @staticmethod
    def _display_datetime(dt: Optional[datetime]) -> str:
        if dt is None:
            return ''
        try:
            return dt.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            return str(dt)

    @staticmethod
    def _berth_duration_text(invoice) -> str:
        det = getattr(invoice, 'detection', None)
        if det is None or det.start_time is None or det.end_time is None:
            return ''
        total_seconds = int((det.end_time - det.start_time).total_seconds())
        if total_seconds <= 0:
            return ''
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f'{hours:02d}:{minutes:02d}:{seconds:02d}'

    @staticmethod
    def _detection_image_link(invoice) -> str:
        det = getattr(invoice, 'detection', None)
        if det is None:
            return ''
        raw = (getattr(det, 'audit_image_path', None) or '').strip()
        if raw:
            return raw
        return ''

    @staticmethod
    def _resolve_image_url(raw_path: str) -> str:
        raw = (raw_path or '').strip()
        if not raw:
            return ''
        if parse_minio_uri(raw) is not None:
            try:
                return presign_get(raw, ttl_seconds=900) or ''
            except Exception:
                return ''
        if raw.startswith('http://') or raw.startswith('https://'):
            return raw
        return ''

    def _embed_invoice_image(
        self,
        ws,
        *,
        row_idx: int,
        col_idx: int,
        invoice,
        temp_files: list[str],
    ) -> None:
        raw_path = self._detection_image_link(invoice)
        if not raw_path:
            ws.cell(row=row_idx, column=col_idx, value='')
            return

        try:
            img_bytes: bytes | None = None
            if parse_minio_uri(raw_path) is not None:
                img_bytes = get_object_bytes(raw_path)
            else:
                img_url = self._resolve_image_url(raw_path)
                if img_url:
                    response = requests.get(img_url, timeout=_IMAGE_FETCH_TIMEOUT_SEC)
                    response.raise_for_status()
                    img_bytes = response.content
            if not img_bytes:
                ws.cell(row=row_idx, column=col_idx, value='')
                return
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            tmp.write(img_bytes)
            tmp.flush()
            tmp.close()
            temp_files.append(tmp.name)

            image = ExcelImage(tmp.name)
            image.width = _IMAGE_THUMBNAIL_SIZE_PX[0]
            image.height = _IMAGE_THUMBNAIL_SIZE_PX[1]
            cell_ref = f'{openpyxl.utils.get_column_letter(col_idx)}{row_idx}'
            ws.add_image(image, cell_ref)
        except Exception as exc:
            _log.warning('Failed to embed invoice image row=%s path=%s err=%s', row_idx, raw_path, exc)
            ws.cell(row=row_idx, column=col_idx, value='')

    @staticmethod
    def _apply_accounting_table_style(ws, header_row: int, data_start_row: int, data_end_row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            header_cell = ws.cell(row=header_row, column=col)
            header_cell.fill = _ACCOUNTING_HEADER_FILL
            header_cell.font = _ACCOUNTING_HEADER_FONT
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = _THIN_BORDER

        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = _THIN_BORDER
                if col in (1, 6, 7):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 10:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    def export_port_logs_excel(
        self,
        db: Session,
        log_date: Optional[date] = None,
        ship_id: Optional[str] = None,
        user_id: Optional[int] = None,
    ) -> str:
        logs = port_log_repo.get_all_logs(db, limit=10000, ship_id=ship_id, log_date=log_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Port Logs'

        headers = [
            'ID',
            'Seq',
            'Ships Completed Today',
            'Logged At',
            'Track ID',
            'Voted Ship ID',
            'First Seen',
            'Last Seen',
            'Confidence',
            'OCR Attempts',
            'Vote Summary (JSON)',
            'Schema Ver',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        for row_idx, log in enumerate(logs, 2):
            vote_json = ''
            if log.vote_summary is not None:
                try:
                    vote_json = json.dumps(log.vote_summary, ensure_ascii=False)
                except (TypeError, ValueError):
                    vote_json = str(log.vote_summary)
            ws.cell(row=row_idx, column=1, value=log.id)
            ws.cell(row=row_idx, column=2, value=log.seq)
            ws.cell(row=row_idx, column=3, value=getattr(log, 'ships_completed_today', None))
            ws.cell(row=row_idx, column=4, value=str(log.logged_at) if log.logged_at else '')
            ws.cell(row=row_idx, column=5, value=log.track_id)
            ws.cell(row=row_idx, column=6, value=log.voted_ship_id)
            ws.cell(row=row_idx, column=7, value=str(log.first_seen_at) if log.first_seen_at else '')
            ws.cell(row=row_idx, column=8, value=str(log.last_seen_at) if log.last_seen_at else '')
            ws.cell(row=row_idx, column=9, value=log.confidence)
            ws.cell(row=row_idx, column=10, value=log.ocr_attempts)
            ws.cell(row=row_idx, column=11, value=vote_json)
            ws.cell(row=row_idx, column=12, value=log.schema_version)

        EXPORTS_DIR.mkdir(exist_ok=True)
        filename = f"port_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = str(EXPORTS_DIR / filename)
        wb.save(file_path)

        # Record the export
        export_log_repo.create(db, {
            'user_id': user_id,
            'export_type': 'port_logs',
            'file_path': file_path,
            'filters': {'date': str(log_date) if log_date else None, 'ship_id': ship_id},
            'row_count': len(logs),
        })

        return file_path

    @staticmethod
    def _payment_status_label(status: Optional[str]) -> str:
        key = (status or 'UNPAID').upper()
        mapping = {
            'PAID': 'Đã thanh toán',
            'PARTIAL': 'Thanh toán một phần',
            'UNPAID': 'Chưa thanh toán',
            'OVERDUE': 'Quá hạn',
            'CANCELLED': 'Đã hủy',
        }
        return mapping.get(key, status or '')

    @staticmethod
    def _berth_minutes(invoice) -> Optional[float]:
        det = getattr(invoice, 'detection', None)
        if det is not None and det.start_time is not None and det.end_time is not None:
            secs = (det.end_time - det.start_time).total_seconds()
            if secs > 0:
                return round(secs / 60, 2)
        return None

    @staticmethod
    def _invoice_ref_fees(invoice) -> str:
        parts: list[str] = []
        for item in invoice.items or []:
            fc = getattr(item, 'fee_config', None)
            if fc is not None and getattr(fc, 'fee_name', None):
                parts.append(str(fc.fee_name))
            elif getattr(item, 'description', None):
                parts.append(str(item.description))
        return '; '.join(parts)

    @staticmethod
    def _invoice_list_kind_label(creation_source: Optional[str]) -> str:
        src = (creation_source or 'USER').upper()
        if src in ('AI', 'ORDER_AUTO'):
            return 'Hóa đơn tự động'
        return 'Hóa đơn tạo tay'

    def _write_revenue_invoices_workbook(self, invoices: list) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bao cao hoa don'
        image_temp_files: list[str] = []

        headers = [
            'STT',
            'MÃ TÀU',
            'TÊN TÀU',
            'LOẠI TÀU',
            'HÌNH ẢNH',
            'SỐ LƯỢT CẬP BẾN',
            'SỐ LƯỢT RỜI BẾN',
            'THỜI GIAN CẬP BẾN',
            'THỜI GIAN NEO ĐẬU',
            'THÀNH TIỀN',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = _ACCOUNTING_HEADER_FILL
            cell.font = _ACCOUNTING_HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = _THIN_BORDER

        for row_idx, inv in enumerate(invoices, 2):
            vessel = inv.vessel
            ship_id = getattr(inv, 'vessel_ship_id_snapshot', None) or (
                vessel.ship_id if vessel else None
            )
            vessel_name = vessel.name if vessel and vessel.name else ''
            vessel_type_name = getattr(inv, 'vessel_type_name_snapshot', None) or (
                vessel.vessel_type.type_name
                if vessel and vessel.vessel_type
                else None
            )
            det = inv.detection
            berth_arrival_count = 1 if det is not None and det.start_time is not None else 0
            berth_departure_count = 1 if det is not None and det.end_time is not None else 0

            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=ship_id or '')
            ws.cell(row=row_idx, column=3, value=vessel_name)
            ws.cell(row=row_idx, column=4, value=vessel_type_name or '')
            self._embed_invoice_image(
                ws,
                row_idx=row_idx,
                col_idx=5,
                invoice=inv,
                temp_files=image_temp_files,
            )
            ws.cell(row=row_idx, column=6, value=berth_arrival_count)
            ws.cell(row=row_idx, column=7, value=berth_departure_count)
            ws.cell(
                row=row_idx,
                column=8,
                value=self._display_datetime(det.start_time) if det is not None else '',
            )
            ws.cell(row=row_idx, column=9, value=self._berth_duration_text(inv))
            amount_cell = ws.cell(row=row_idx, column=10, value=float(inv.total_amount or 0))
            amount_cell.number_format = '#,##0'
            ws.row_dimensions[row_idx].height = 72

        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 24
        widths = {
            1: 8,
            2: 18,
            3: 24,
            4: 16,
            5: 42,
            6: 16,
            7: 16,
            8: 22,
            9: 18,
            10: 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        self._apply_accounting_table_style(
            ws,
            header_row=1,
            data_start_row=2,
            data_end_row=max(2, len(invoices) + 1),
            col_count=len(headers),
        )
        wb._image_temp_files = image_temp_files  # type: ignore[attr-defined]
        return wb

    @staticmethod
    def _cleanup_workbook_temp_images(wb: openpyxl.Workbook) -> None:
        files = getattr(wb, '_image_temp_files', [])
        for file_path in files:
            try:
                Path(file_path).unlink(missing_ok=True)
            except Exception:
                continue

    def export_revenue_invoices_excel(
        self,
        db: Session,
        invoice_ids: list[int],
        *,
        user_id: Optional[int] = None,
        list_kind: Optional[str] = None,
        invoice_sub_tab: Optional[str] = None,
        filters: Optional[dict] = None,
    ) -> str:
        invoices = invoice_repo.get_by_ids(db, invoice_ids, include_deleted=True)
        if not invoices:
            raise ValueError('No invoices found for export')

        wb = self._write_revenue_invoices_workbook(invoices)

        EXPORTS_DIR.mkdir(exist_ok=True)
        kind = (list_kind or 'invoices').replace('_', '-')
        tab = (invoice_sub_tab or 'all').replace('_', '-')
        filename = f"revenue_{kind}_{tab}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = str(EXPORTS_DIR / filename)
        try:
            wb.save(file_path)
        finally:
            self._cleanup_workbook_temp_images(wb)

        export_log_repo.create(
            db,
            {
                'user_id': user_id,
                'export_type': 'revenue_invoices',
                'file_path': file_path,
                'filters': filters or {'invoice_ids': invoice_ids, 'list_kind': list_kind},
                'row_count': len(invoices),
            },
        )
        return file_path

    def export_all_revenue_invoices_excel(
        self,
        db: Session,
        *,
        user_id: Optional[int] = None,
    ) -> str:
        invoices = invoice_repo.get_all_revenue_export(db)
        if not invoices:
            raise ValueError('No invoices found for export')

        wb = self._write_revenue_invoices_workbook(invoices)

        EXPORTS_DIR.mkdir(exist_ok=True)
        filename = f"revenue_all_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = str(EXPORTS_DIR / filename)
        try:
            wb.save(file_path)
        finally:
            self._cleanup_workbook_temp_images(wb)

        export_log_repo.create(
            db,
            {
                'user_id': user_id,
                'export_type': 'revenue_invoices_all',
                'file_path': file_path,
                'filters': {'scope': 'all_revenue_history'},
                'row_count': len(invoices),
            },
        )
        return file_path

    def export_revenue_fee_configs_excel(
        self,
        db: Session,
        fee_config_ids: list[int],
        *,
        user_id: Optional[int] = None,
        filters: Optional[dict] = None,
    ) -> str:
        fees = fee_config_repo.get_by_ids(db, fee_config_ids)
        if not fees:
            raise ValueError('No fee configs found for export')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cấu hình phí'

        headers = [
            'ID',
            'Tên phí',
            'Loại tàu',
            'Đơn giá',
            'Đơn vị',
            'Đang áp dụng',
            'Giới hạn neo',
            'Hiệu lực từ',
            'Hiệu lực đến',
            'Tạo lúc',
            'Cập nhật',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        for row_idx, fee in enumerate(fees, 2):
            limit_label = ''
            if fee.berth_limit_count and fee.berth_limit_unit:
                unit = 'ngày' if fee.berth_limit_unit == 'day' else 'tháng'
                limit_label = f'{fee.berth_limit_count} lần/{unit}'
            ws.cell(row=row_idx, column=1, value=fee.id)
            ws.cell(row=row_idx, column=2, value=fee.fee_name)
            ws.cell(
                row=row_idx,
                column=3,
                value=fee.vessel_type.type_name if fee.vessel_type else '',
            )
            ws.cell(row=row_idx, column=4, value=float(fee.base_fee or 0))
            ws.cell(row=row_idx, column=5, value=fee_billing_unit_label(fee.unit))
            ws.cell(row=row_idx, column=6, value='Có' if fee.is_active else 'Không')
            ws.cell(row=row_idx, column=7, value=limit_label)
            ws.cell(
                row=row_idx,
                column=8,
                value=str(fee.effective_from) if fee.effective_from else '',
            )
            ws.cell(
                row=row_idx,
                column=9,
                value=str(fee.effective_to) if fee.effective_to else '',
            )
            ws.cell(row=row_idx, column=10, value=str(fee.created_at) if fee.created_at else '')
            ws.cell(row=row_idx, column=11, value=str(fee.updated_at) if fee.updated_at else '')

        EXPORTS_DIR.mkdir(exist_ok=True)
        filename = f"revenue_fee_configs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = str(EXPORTS_DIR / filename)
        wb.save(file_path)

        export_log_repo.create(
            db,
            {
                'user_id': user_id,
                'export_type': 'revenue_fee_configs',
                'file_path': file_path,
                'filters': filters or {'fee_config_ids': fee_config_ids},
                'row_count': len(fees),
            },
        )
        return file_path


export_service = ExportService()
